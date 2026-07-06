"""
OrbitFlow -- full project management web application backed entirely by
OrbitFlow_PM_Tracker_v3.xlsx (no SQL/NoSQL database of any kind).

RELATIONAL MODEL (unchanged from the data-entry app this replaces):
  SOURCE sheets (the only sheets ever written to):
    - Projects Master
    - Resources
    - Work Items            (Type = Task | Bug | Risk | CR)

  VIEW / REPORTING sheets (never written to by this app):
    - Risk Register          = Work Items filtered to Type="Risk"
    - Change Requests        = Work Items filtered to Type="CR"
    - Project Dashboard, Financial Tracker, Resource Cost, CONFIG's
      project registry, Billing Milestones' summary table -- all
      pre-built formulas that pull from the source sheets automatically.

HEADER DETECTION: column positions for every field are resolved at
request time by reading each sheet's header row and matching on the
header text (see `headers_of`), rather than hardcoded column numbers --
new columns can be appended to a source sheet without any code changes.

DELETE SAFETY: Financial Tracker / Project Dashboard address Projects
Master positionally via ROW()-relative formulas, so removing a row and
shifting everything up is safe there. But CONFIG's project registry,
Billing Milestones' summary table, and Resource Cost's project
cost-benefit block address Projects Master with a HARDCODED row number
per formula (baked in when those sheets were built) -- shifting rows
would silently misalign them. So Projects Master / Resources rows are
never physically removed; "Delete" blanks the row in place (position
preserved) and "Archive" just flips Status. Work Items has no such
hardcoded-row dependents (everything reads it via whole-column
SUMIFS/COUNTIFS/array-extraction), so work items ARE deleted for real.

Run:
    pip install flask openpyxl
    python server.py
Then open http://127.0.0.1:5000
"""
import os
import re
import io
import copy
import datetime
from pathlib import Path
from collections import Counter

from flask import Flask, request, jsonify, send_from_directory, send_file
import openpyxl
from openpyxl.utils import get_column_letter

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_XLSX = BASE_DIR / "OrbitFlow_PM_Tracker_v3.xlsx"
XLSX_PATH = Path(os.environ.get("XLSX_PATH", DEFAULT_XLSX))

# On a host with a persistent disk (e.g. Render), XLSX_PATH points at that
# disk's mount, which starts out empty on first deploy. Seed it from the
# workbook bundled in the repo so the very first boot has real data instead
# of a missing-file error; every boot after that finds the file already
# there and leaves it alone.
if not XLSX_PATH.exists() and DEFAULT_XLSX.exists() and XLSX_PATH != DEFAULT_XLSX:
    XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    XLSX_PATH.write_bytes(DEFAULT_XLSX.read_bytes())

app = Flask(__name__, static_folder=str(BASE_DIR / "static"), static_url_path="")

SHEET_PM = "Projects Master"
SHEET_RES = "Resources"
SHEET_WI = "Work Items"

PM_HEADER_ROW = 4
PM_FIRST_DATA_ROW = 5
RES_HEADER_ROW = 6
RES_FIRST_DATA_ROW = 7
WI_HEADER_ROW = 5
WI_FIRST_DATA_ROW = 6


# ---------------------------------------------------------------------------
# workbook + header-detection helpers
# ---------------------------------------------------------------------------

_wb_cache = {"wb": None, "mtime": None}


def load_wb():
    """Cache the parsed workbook in memory per-process, keyed on the file's
    mtime. openpyxl's load_workbook() re-parses every sheet/style/formula
    from scratch, which is the main cost on a slow/shared-CPU host -- this
    avoids paying that cost on every single GET request. Any write goes
    through save_wb() below, which keeps the cache in sync with what's on
    disk instead of invalidating it, so the very next request doesn't
    trigger a redundant re-read of what we just wrote ourselves."""
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"Could not find {XLSX_PATH}.")
    mtime = XLSX_PATH.stat().st_mtime
    if _wb_cache["wb"] is not None and _wb_cache["mtime"] == mtime:
        return _wb_cache["wb"]
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=False)
    _wb_cache["wb"] = wb
    _wb_cache["mtime"] = mtime
    return wb


def save_wb(wb):
    wb.save(XLSX_PATH)
    _wb_cache["wb"] = wb
    _wb_cache["mtime"] = XLSX_PATH.stat().st_mtime


def headers_of(ws, header_row):
    """{header text: 1-based column index} for a sheet, read live -- so
    appended columns are picked up automatically, no hardcoded indices."""
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v:
            out[str(v).strip()] = c
    return out


def h(headers, *names, default=None):
    """First matching header column index out of several accepted spellings."""
    for n in names:
        if n in headers:
            return headers[n]
    return default


def last_data_row(ws, first_row, id_col=1, skip_prefix=None):
    last = first_row - 1
    for r in range(first_row, ws.max_row + 1):
        v = ws.cell(row=r, column=id_col).value
        if v and not (skip_prefix and str(v).lower().startswith(skip_prefix)):
            last = r
    return last


def parse_date(val):
    if not val:
        return None
    if isinstance(val, (datetime.date, datetime.datetime)):
        return val
    try:
        return datetime.datetime.strptime(str(val), "%Y-%m-%d")
    except ValueError:
        return val


def cell_json(v):
    if isinstance(v, (datetime.date, datetime.datetime)):
        return v.strftime("%Y-%m-%d")
    return v


def _shift_formula_rows(formula, old_row, new_row):
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    def repl(m):
        colpart, row = m.group(1), m.group(2)
        return f"{colpart}{new_row}" if int(row) == old_row else m.group(0)

    return re.sub(r"(\$?[A-Za-z]{1,2})(\d+)", repl, formula)


def next_seq_id(ws, id_col, prefix):
    max_n = 0
    for row in ws.iter_rows(min_row=1, max_col=id_col, values_only=False):
        cell = row[id_col - 1]
        if cell.value and isinstance(cell.value, str) and cell.value.startswith(prefix):
            m = re.search(r"(\d+)$", cell.value)
            if m:
                max_n = max(max_n, int(m.group(1)))
    return f"{prefix}{max_n + 1:03d}"


def project_code(project_name):
    words = re.findall(r"[A-Za-z0-9]+", project_name or "")
    return "".join(w[0] for w in words[:3]).upper() or "GEN"


TYPE_PREFIX = {"Task": "T", "Bug": "B", "Risk": "RK", "CR": "CR"}


# ---------------------------------------------------------------------------
# row <-> dict helpers (generic, header-driven)
# ---------------------------------------------------------------------------

def row_to_dict(ws, row, headers, field_map=None):
    """field_map: {json_key: header_text} -- if omitted, uses header text as key."""
    out = {}
    field_map = field_map or {hh: hh for hh in headers}
    for key, header_text in field_map.items():
        col = headers.get(header_text)
        if col:
            out[key] = cell_json(ws.cell(row=row, column=col).value)
    out["_row"] = row
    return out


def write_fields(ws, row, headers, data, field_map, date_fields=()):
    for key, header_text in field_map.items():
        if key not in data:
            continue
        col = headers.get(header_text)
        if not col:
            continue
        val = data[key]
        if key in date_fields:
            val = parse_date(val)
        cell = ws.cell(row=row, column=col, value=val)
        if key in date_fields:
            cell.number_format = "yyyy-mm-dd"


PM_FIELDS = {
    "project_id": "Project ID", "project_name": "Project Name", "pm": "PM",
    "account_manager": "Account Manager", "project_type": "Project Type",
    "status": "Status", "start_date": "Start Date", "end_date": "End Date",
    "sow_hrs": "SOW Hrs", "domain": "Domain", "sow_value": "SOW Value ($)",
    "priority": "Priority", "health": "Health", "business_unit": "Business Unit",
    "delivery_manager": "Delivery Manager", "technology": "Technology",
    "currency": "Currency", "description": "Description",
}
PM_DATES = {"start_date", "end_date"}

RES_FIELDS = {
    "employee": "Employee", "level": "Level", "manager": "Manager", "role": "Role",
    "resource_type": "Resource Type", "onsite_offshore": "Onsite/Offshore",
    "email": "Email", "designation": "Designation", "availability": "Availability %",
    "billing_rate": "Billing Rate ($/hr)", "joining_date": "Joining Date",
    "skills": "Skills", "experience": "Experience (yrs)",
}
RES_DATES = {"joining_date"}

WI_FIELDS = {
    "type": "Type", "id": "ID", "summary": "Summary", "module": "Module",
    "submodule": "SubModule", "event_bug_type": "Event / BugType", "category": "Category",
    "activity": "Activity", "sub_activity": "Sub-Activity", "assigned_to": "Assigned To",
    "priority": "Priority", "status": "Status", "worked_date": "Worked Date",
    "original_effort": "Original Effort (hrs)", "actual_effort": "Actual Effort (hrs)",
    "start_date": "Start Date", "end_date": "End Date", "probability": "Probability (1-5)",
    "impact": "Impact (1-5)", "mitigation_hrs": "Mitigation Hrs",
    "description": "Description", "story_points": "Story Points", "remarks": "Remarks",
    "completed_date": "Completed Date", "legacy_risk_score": "Legacy Risk Score",
}
WI_DATES = {"worked_date", "start_date", "end_date", "completed_date"}


def risk_score(item):
    """Probability x Impact when both are set; otherwise fall back to the
    Legacy Risk Score column (older risks recorded before this app tracked
    Probability/Impact separately -- see migration notes)."""
    p, i = item.get("probability"), item.get("impact")
    if p and i:
        return p * i
    return item.get("legacy_risk_score") or 0


# ---------------------------------------------------------------------------
# DASHBOARD
# ---------------------------------------------------------------------------

@app.route("/api/dashboard")
def dashboard():
    wb = load_wb()
    pm = wb[SHEET_PM]
    res = wb[SHEET_RES]
    wi = wb[SHEET_WI]
    pmh = headers_of(pm, PM_HEADER_ROW)
    resh = headers_of(res, RES_HEADER_ROW)
    wih = headers_of(wi, WI_HEADER_ROW)

    projects = []
    for r in range(PM_FIRST_DATA_ROW, last_data_row(pm, PM_FIRST_DATA_ROW) + 1):
        projects.append(row_to_dict(pm, r, pmh, PM_FIELDS))

    resources = []
    for r in range(RES_FIRST_DATA_ROW, last_data_row(res, RES_FIRST_DATA_ROW, skip_prefix="note") + 1):
        emp = res.cell(row=r, column=1).value
        if emp:
            resources.append(row_to_dict(res, r, resh, RES_FIELDS))

    items = []
    for r in range(WI_FIRST_DATA_ROW, last_data_row(wi, WI_FIRST_DATA_ROW) + 1):
        items.append(row_to_dict(wi, r, wih, WI_FIELDS))

    total_projects = len(projects)
    active_projects = sum(1 for p in projects if p.get("status") == "Active")
    completed_projects = sum(1 for p in projects if p.get("status") in ("Closed", "Completed"))
    delayed_projects = sum(1 for p in projects if p.get("health") in ("Red", "Delayed"))

    tasks = [i for i in items if i.get("type") == "Task"]
    bugs = [i for i in items if i.get("type") == "Bug"]
    risks = [i for i in items if i.get("type") == "Risk"]
    crs = [i for i in items if i.get("type") == "CR"]

    def is_open(i):
        return i.get("status") not in ("Completed", "Fixed", "Closed", "Rejected", "Approved")

    open_tasks = sum(1 for i in tasks if is_open(i))
    completed_tasks = len(tasks) - open_tasks
    open_bugs = sum(1 for i in bugs if is_open(i))
    closed_bugs = len(bugs) - open_bugs
    open_risks = sum(1 for i in risks if is_open(i))
    high_risks = sum(1 for i in risks if risk_score(i) >= 15)
    open_crs = sum(1 for i in crs if is_open(i))

    planned_hours = sum((i.get("original_effort") or 0) for i in items)
    actual_hours = sum((i.get("actual_effort") or 0) for i in items)
    total_sow_hrs = sum((p.get("sow_hrs") or 0) for p in projects)
    budget_utilization = round((actual_hours / total_sow_hrs) * 100, 1) if total_sow_hrs else 0

    status_counts = Counter(i.get("status") or "Unknown" for i in tasks + bugs)
    risk_severity = Counter()
    for rk in risks:
        score = risk_score(rk)
        band = "Critical" if score >= 15 else "High" if score >= 10 else "Medium" if score >= 5 else "Low"
        risk_severity[band] += 1
    health_counts = Counter(p.get("health") or "Unknown" for p in projects)

    workload = Counter()
    for i in tasks + bugs:
        if i.get("assigned_to"):
            workload[i["assigned_to"]] += (i.get("actual_effort") or 0)

    upcoming_deadlines = sorted(
        [i for i in items if i.get("end_date") and is_open(i)],
        key=lambda i: i["end_date"],
    )[:8]

    recent_activity = sorted(
        [i for i in items if i.get("worked_date")],
        key=lambda i: i["worked_date"], reverse=True,
    )[:10]

    pending_approvals = [i for i in crs if i.get("status") == "Pending Approval"]

    return jsonify({
        "stats": {
            "total_projects": total_projects, "active_projects": active_projects,
            "completed_projects": completed_projects, "delayed_projects": delayed_projects,
            "total_resources": len(resources), "open_tasks": open_tasks,
            "completed_tasks": completed_tasks, "open_bugs": open_bugs, "closed_bugs": closed_bugs,
            "open_risks": open_risks, "high_risks": high_risks, "open_crs": open_crs,
            "budget_utilization": budget_utilization, "planned_hours": planned_hours,
            "actual_hours": actual_hours,
        },
        "charts": {
            "project_health": dict(health_counts),
            "tasks_by_status": dict(status_counts),
            "risks_by_severity": dict(risk_severity),
            "workload": dict(workload.most_common(10)),
            "budget_utilization": budget_utilization,
        },
        "upcoming_deadlines": upcoming_deadlines,
        "recent_activity": recent_activity,
        "pending_approvals": pending_approvals,
    })


# ---------------------------------------------------------------------------
# PROJECTS
# ---------------------------------------------------------------------------

@app.route("/api/projects", methods=["GET"])
def list_projects():
    wb = load_wb()
    pm = wb[SHEET_PM]
    pmh = headers_of(pm, PM_HEADER_ROW)
    q = (request.args.get("q") or "").lower()
    status = request.args.get("status")
    priority = request.args.get("priority")
    health = request.args.get("health")
    domain = request.args.get("domain")
    sort = request.args.get("sort")
    order = request.args.get("order", "asc")
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 25))

    rows = []
    for r in range(PM_FIRST_DATA_ROW, last_data_row(pm, PM_FIRST_DATA_ROW) + 1):
        d = row_to_dict(pm, r, pmh, PM_FIELDS)
        if not d.get("project_id"):
            continue
        rows.append(d)

    if q:
        rows = [d for d in rows if q in " ".join(str(v).lower() for v in d.values() if v)]
    if status:
        rows = [d for d in rows if d.get("status") == status]
    if priority:
        rows = [d for d in rows if d.get("priority") == priority]
    if health:
        rows = [d for d in rows if d.get("health") == health]
    if domain:
        rows = [d for d in rows if d.get("domain") == domain]
    if sort:
        rows.sort(key=lambda d: (d.get(sort) is None, d.get(sort)), reverse=(order == "desc"))

    total = len(rows)
    start = (page - 1) * page_size
    page_rows = rows[start:start + page_size]
    return jsonify({"total": total, "page": page, "page_size": page_size, "items": page_rows})


@app.route("/api/projects/<project_id>", methods=["GET"])
def get_project(project_id):
    wb = load_wb()
    pm, res, wi = wb[SHEET_PM], wb[SHEET_RES], wb[SHEET_WI]
    pmh, resh, wih = headers_of(pm, PM_HEADER_ROW), headers_of(res, RES_HEADER_ROW), headers_of(wi, WI_HEADER_ROW)

    project = None
    for r in range(PM_FIRST_DATA_ROW, last_data_row(pm, PM_FIRST_DATA_ROW) + 1):
        d = row_to_dict(pm, r, pmh, PM_FIELDS)
        if d.get("project_id") == project_id:
            project = d
            break
    if not project:
        return jsonify({"error": "Project not found"}), 404

    project_name = project.get("project_name")
    items = []
    for r in range(WI_FIRST_DATA_ROW, last_data_row(wi, WI_FIRST_DATA_ROW) + 1):
        d = row_to_dict(wi, r, wih, WI_FIELDS)
        if d.get("module") == project_name:
            items.append(d)

    assigned_names = {i.get("assigned_to") for i in items if i.get("assigned_to")}
    resources = []
    for r in range(RES_FIRST_DATA_ROW, last_data_row(res, RES_FIRST_DATA_ROW, skip_prefix="note") + 1):
        emp = res.cell(row=r, column=1).value
        if emp and emp in assigned_names:
            resources.append(with_utilization(row_to_dict(res, r, resh, RES_FIELDS)))

    work_items = _all_work_items(wb)
    financial = _financial_row(project, work_items)
    billing = _billing_row(wb, project)

    return jsonify({
        "project": project,
        "resources": resources,
        "tasks": [i for i in items if i["type"] == "Task"],
        "bugs": [i for i in items if i["type"] == "Bug"],
        "risks": [i for i in items if i["type"] == "Risk"],
        "change_requests": [i for i in items if i["type"] == "CR"],
        "financial": financial,
        "billing": billing,
    })


@app.route("/api/projects", methods=["POST"])
def create_project():
    d = request.json
    wb = load_wb()
    pm = wb[SHEET_PM]
    pmh = headers_of(pm, PM_HEADER_ROW)
    new_row = last_data_row(pm, PM_FIRST_DATA_ROW) + 1
    project_id = next_seq_id(pm, pmh.get("Project ID", 1), "P")
    d = {**d, "project_id": project_id, "status": d.get("status", "Active"),
         "priority": d.get("priority", "Medium"), "health": d.get("health", "Green"),
         "currency": d.get("currency", "USD")}
    write_fields(pm, new_row, pmh, d, PM_FIELDS, PM_DATES)
    save_wb(wb)
    return jsonify({"ok": True, "project_id": project_id})


@app.route("/api/projects/<project_id>", methods=["PUT"])
def update_project(project_id):
    d = request.json
    wb = load_wb()
    pm = wb[SHEET_PM]
    pmh = headers_of(pm, PM_HEADER_ROW)
    id_col = pmh.get("Project ID", 1)
    target_row = None
    for r in range(PM_FIRST_DATA_ROW, last_data_row(pm, PM_FIRST_DATA_ROW) + 1):
        if pm.cell(row=r, column=id_col).value == project_id:
            target_row = r
            break
    if not target_row:
        return jsonify({"error": "Project not found"}), 404
    write_fields(pm, target_row, pmh, d, PM_FIELDS, PM_DATES)
    save_wb(wb)
    return jsonify({"ok": True})


@app.route("/api/projects/<project_id>", methods=["DELETE"])
def delete_project(project_id):
    """Blanks the row in place rather than removing it -- CONFIG's project
    registry, Billing Milestones' summary, and Resource Cost's project
    cost-benefit block all address Projects Master rows by a hardcoded
    row number, so shifting rows up would silently misalign them."""
    wb = load_wb()
    pm = wb[SHEET_PM]
    pmh = headers_of(pm, PM_HEADER_ROW)
    id_col = pmh.get("Project ID", 1)
    target_row = None
    for r in range(PM_FIRST_DATA_ROW, last_data_row(pm, PM_FIRST_DATA_ROW) + 1):
        if pm.cell(row=r, column=id_col).value == project_id:
            target_row = r
            break
    if not target_row:
        return jsonify({"error": "Project not found"}), 404
    for c in range(1, pm.max_column + 1):
        pm.cell(row=target_row, column=c).value = None
    save_wb(wb)
    return jsonify({"ok": True})


@app.route("/api/projects/<project_id>/archive", methods=["POST"])
def archive_project(project_id):
    wb = load_wb()
    pm = wb[SHEET_PM]
    pmh = headers_of(pm, PM_HEADER_ROW)
    id_col = pmh.get("Project ID", 1)
    status_col = pmh.get("Status")
    for r in range(PM_FIRST_DATA_ROW, last_data_row(pm, PM_FIRST_DATA_ROW) + 1):
        if pm.cell(row=r, column=id_col).value == project_id:
            pm.cell(row=r, column=status_col, value="Archived")
            save_wb(wb)
            return jsonify({"ok": True})
    return jsonify({"error": "Project not found"}), 404


@app.route("/api/projects/<project_id>/duplicate", methods=["POST"])
def duplicate_project(project_id):
    wb = load_wb()
    pm = wb[SHEET_PM]
    pmh = headers_of(pm, PM_HEADER_ROW)
    id_col = pmh.get("Project ID", 1)
    src_row = None
    for r in range(PM_FIRST_DATA_ROW, last_data_row(pm, PM_FIRST_DATA_ROW) + 1):
        if pm.cell(row=r, column=id_col).value == project_id:
            src_row = r
            break
    if not src_row:
        return jsonify({"error": "Project not found"}), 404
    new_row = last_data_row(pm, PM_FIRST_DATA_ROW) + 1
    new_id = next_seq_id(pm, id_col, "P")
    name_col = pmh.get("Project Name")
    for c in range(1, pm.max_column + 1):
        pm.cell(row=new_row, column=c, value=pm.cell(row=src_row, column=c).value)
    pm.cell(row=new_row, column=id_col, value=new_id)
    if name_col:
        orig_name = pm.cell(row=src_row, column=name_col).value
        pm.cell(row=new_row, column=name_col, value=f"{orig_name} (Copy)")
    save_wb(wb)
    return jsonify({"ok": True, "project_id": new_id})


# ---------------------------------------------------------------------------
# RESOURCES
# ---------------------------------------------------------------------------

FTE_MONTHLY_HOURS = 160  # 1 full-time resource = 160 working hours/month


def with_utilization(r):
    """Monthly allocation math: Availability % is the resource's assigned
    allocation on this project workload (100% = 160 hrs, 75% = 120, etc).
    Purely derived -- not stored in Excel, since it's just availability x
    160 restated a few different ways for the UI."""
    availability = r.get("availability")
    availability = 100 if availability is None else availability
    allocated_hours = round(FTE_MONTHLY_HOURS * (availability / 100), 1)
    available_hours = FTE_MONTHLY_HOURS
    remaining_capacity = round(available_hours - allocated_hours, 1)
    r["allocated_hours"] = allocated_hours
    r["available_hours"] = available_hours
    r["remaining_capacity"] = remaining_capacity
    r["utilization_pct"] = round((allocated_hours / available_hours) * 100, 1) if available_hours else 0
    r["overallocated"] = allocated_hours > available_hours
    return r


@app.route("/api/resources", methods=["GET"])
def list_resources():
    wb = load_wb()
    res = wb[SHEET_RES]
    resh = headers_of(res, RES_HEADER_ROW)
    q = (request.args.get("q") or "").lower()
    role = request.args.get("role")
    dept = request.args.get("department") or request.args.get("resource_type")
    sort = request.args.get("sort")
    order = request.args.get("order", "asc")

    rows = []
    for r in range(RES_FIRST_DATA_ROW, last_data_row(res, RES_FIRST_DATA_ROW, skip_prefix="note") + 1):
        emp = res.cell(row=r, column=1).value
        if not emp:
            continue
        rows.append(with_utilization(row_to_dict(res, r, resh, RES_FIELDS)))

    if q:
        rows = [d for d in rows if q in " ".join(str(v).lower() for v in d.values() if v)]
    if role:
        rows = [d for d in rows if d.get("role") == role]
    if dept:
        rows = [d for d in rows if d.get("resource_type") == dept]
    if sort:
        rows.sort(key=lambda d: (d.get(sort) is None, d.get(sort)), reverse=(order == "desc"))

    return jsonify({"total": len(rows), "items": rows})


def _check_allocation(d):
    """Availability % above 100 means the resource is assigned more hours
    than a single FTE has in a month (160 hrs). Block it unless the caller
    explicitly passes override_allocation -- e.g. a deliberate short-term
    crunch decision, not an accidental typo."""
    availability = d.get("availability")
    if availability is not None and availability > 100 and not d.get("override_allocation"):
        return jsonify({
            "error": "over_allocated",
            "message": f"{availability}% allocation exceeds available capacity (160 hrs/month at 100%). "
                       "Resubmit with override_allocation: true to proceed anyway.",
        }), 409
    return None


@app.route("/api/resources", methods=["POST"])
def create_resource():
    d = request.json
    err = _check_allocation(d)
    if err:
        return err
    wb = load_wb()
    ws = wb[SHEET_RES]
    resh = headers_of(ws, RES_HEADER_ROW)
    last_row = last_data_row(ws, RES_FIRST_DATA_ROW, skip_prefix="note")
    new_row = last_row + 1
    ws.insert_rows(new_row)
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=last_row, column=col)
        dst = ws.cell(row=new_row, column=col)
        dst._style = copy.copy(src._style)
        if isinstance(src.value, str) and src.value.startswith("="):
            dst.value = _shift_formula_rows(src.value, last_row, new_row)

    is_support = bool(d.get("is_support_team"))
    working_cap_col = resh.get("Working Cap (hrs)")
    if working_cap_col:
        ws.cell(row=new_row, column=working_cap_col, value=80 if is_support else "=$M$11")

    write_fields(ws, new_row, resh, d, RES_FIELDS, RES_DATES)
    if resh.get("Availability %") and "availability" not in d:
        ws.cell(row=new_row, column=resh["Availability %"], value=100)
    save_wb(wb)
    return jsonify({"ok": True, "row": new_row})


@app.route("/api/resources/<employee>", methods=["PUT"])
def update_resource(employee):
    d = request.json
    err = _check_allocation(d)
    if err:
        return err
    wb = load_wb()
    ws = wb[SHEET_RES]
    resh = headers_of(ws, RES_HEADER_ROW)
    target_row = None
    for r in range(RES_FIRST_DATA_ROW, last_data_row(ws, RES_FIRST_DATA_ROW, skip_prefix="note") + 1):
        if ws.cell(row=r, column=1).value == employee:
            target_row = r
            break
    if not target_row:
        return jsonify({"error": "Resource not found"}), 404
    write_fields(ws, target_row, resh, d, RES_FIELDS, RES_DATES)
    save_wb(wb)
    return jsonify({"ok": True})


@app.route("/api/resources/<employee>", methods=["DELETE"])
def delete_resource(employee):
    """Blanked in place, same reasoning as Delete Project."""
    wb = load_wb()
    ws = wb[SHEET_RES]
    target_row = None
    for r in range(RES_FIRST_DATA_ROW, last_data_row(ws, RES_FIRST_DATA_ROW, skip_prefix="note") + 1):
        if ws.cell(row=r, column=1).value == employee:
            target_row = r
            break
    if not target_row:
        return jsonify({"error": "Resource not found"}), 404
    for c in range(1, ws.max_column + 1):
        ws.cell(row=target_row, column=c).value = None
    save_wb(wb)
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# WORK ITEMS (Tasks / Bugs / Risks / CRs -- one unified sheet)
# ---------------------------------------------------------------------------

@app.route("/api/workitems", methods=["GET"])
def list_workitems():
    wb = load_wb()
    wi = wb[SHEET_WI]
    wih = headers_of(wi, WI_HEADER_ROW)
    q = (request.args.get("q") or "").lower()
    wtype = request.args.get("type")
    status = request.args.get("status")
    priority = request.args.get("priority")
    project = request.args.get("project")
    assigned_to = request.args.get("assigned_to")
    sort = request.args.get("sort")
    order = request.args.get("order", "asc")
    page = int(request.args.get("page", 1))
    page_size = int(request.args.get("page_size", 50))

    rows = []
    for r in range(WI_FIRST_DATA_ROW, last_data_row(wi, WI_FIRST_DATA_ROW) + 1):
        d = row_to_dict(wi, r, wih, WI_FIELDS)
        if d.get("id"):
            rows.append(d)

    if wtype:
        rows = [d for d in rows if d.get("type") == wtype]
    if status:
        rows = [d for d in rows if d.get("status") == status]
    if priority:
        rows = [d for d in rows if d.get("priority") == priority]
    if project:
        rows = [d for d in rows if d.get("module") == project]
    if assigned_to:
        rows = [d for d in rows if d.get("assigned_to") == assigned_to]
    if q:
        rows = [d for d in rows if q in " ".join(str(v).lower() for v in d.values() if v)]
    if sort:
        rows.sort(key=lambda d: (d.get(sort) is None, d.get(sort)), reverse=(order == "desc"))

    total = len(rows)
    start = (page - 1) * page_size
    page_rows = rows[start:start + page_size]
    return jsonify({"total": total, "page": page, "page_size": page_size, "items": page_rows})


@app.route("/api/workitems", methods=["POST"])
def create_workitem():
    d = request.json
    wb = load_wb()
    ws = wb[SHEET_WI]
    wih = headers_of(ws, WI_HEADER_ROW)
    wtype = d.get("type")
    new_row = last_data_row(ws, WI_FIRST_DATA_ROW) + 1
    code = project_code(d.get("module"))
    item_id = next_seq_id(ws, wih.get("ID", 2), f"{code}-{TYPE_PREFIX.get(wtype, 'X')}")
    d = {**d, "id": item_id}
    write_fields(ws, new_row, wih, d, WI_FIELDS, WI_DATES)
    save_wb(wb)
    return jsonify({"ok": True, "id": item_id})


@app.route("/api/workitems/<item_id>", methods=["PUT"])
def update_workitem(item_id):
    d = request.json
    wb = load_wb()
    ws = wb[SHEET_WI]
    wih = headers_of(ws, WI_HEADER_ROW)
    id_col = wih.get("ID", 2)
    target_row = None
    for r in range(WI_FIRST_DATA_ROW, last_data_row(ws, WI_FIRST_DATA_ROW) + 1):
        if ws.cell(row=r, column=id_col).value == item_id:
            target_row = r
            break
    if not target_row:
        return jsonify({"error": "Work item not found"}), 404
    write_fields(ws, target_row, wih, d, WI_FIELDS, WI_DATES)
    save_wb(wb)
    return jsonify({"ok": True})


@app.route("/api/workitems/<item_id>", methods=["DELETE"])
def delete_workitem(item_id):
    """Real row deletion -- every consumer of Work Items reads it via
    whole-column ranges (SUMIFS/COUNTIFS/INDEX+SMALL array extraction),
    not hardcoded row numbers, so shifting rows up is safe here."""
    wb = load_wb()
    ws = wb[SHEET_WI]
    wih = headers_of(ws, WI_HEADER_ROW)
    id_col = wih.get("ID", 2)
    target_row = None
    for r in range(WI_FIRST_DATA_ROW, last_data_row(ws, WI_FIRST_DATA_ROW) + 1):
        if ws.cell(row=r, column=id_col).value == item_id:
            target_row = r
            break
    if not target_row:
        return jsonify({"error": "Work item not found"}), 404
    ws.delete_rows(target_row, 1)
    save_wb(wb)
    return jsonify({"ok": True})


@app.route("/api/workitems/bulk", methods=["PATCH"])
def bulk_update_workitems():
    d = request.json
    ids = d.get("ids", [])
    changes = d.get("changes", {})
    wb = load_wb()
    ws = wb[SHEET_WI]
    wih = headers_of(ws, WI_HEADER_ROW)
    id_col = wih.get("ID", 2)
    updated = 0
    for r in range(WI_FIRST_DATA_ROW, last_data_row(ws, WI_FIRST_DATA_ROW) + 1):
        if ws.cell(row=r, column=id_col).value in ids:
            write_fields(ws, r, wih, changes, WI_FIELDS, WI_DATES)
            updated += 1
    save_wb(wb)
    return jsonify({"ok": True, "updated": updated})


# ---------------------------------------------------------------------------
# RISKS / CHANGE REQUESTS -- read-only filtered views (Type=Risk / Type=CR)
# ---------------------------------------------------------------------------

@app.route("/api/risks", methods=["GET"])
def list_risks():
    request.args = request.args.copy()
    with app.test_request_context(f"/api/workitems?{request.query_string.decode()}&type=Risk"):
        return list_workitems()


@app.route("/api/change-requests", methods=["GET"])
def list_change_requests():
    with app.test_request_context(f"/api/workitems?{request.query_string.decode()}&type=CR"):
        return list_workitems()


def _set_workitem_status(item_id, status, expected_type=None):
    wb = load_wb()
    ws = wb[SHEET_WI]
    wih = headers_of(ws, WI_HEADER_ROW)
    id_col = wih.get("ID", 2)
    type_col = wih.get("Type", 1)
    status_col = wih.get("Status")
    for r in range(WI_FIRST_DATA_ROW, last_data_row(ws, WI_FIRST_DATA_ROW) + 1):
        if ws.cell(row=r, column=id_col).value == item_id:
            if expected_type and ws.cell(row=r, column=type_col).value != expected_type:
                return jsonify({"error": f"{item_id} is not a {expected_type}"}), 400
            ws.cell(row=r, column=status_col, value=status)
            save_wb(wb)
            return jsonify({"ok": True})
    return jsonify({"error": "Not found"}), 404


@app.route("/api/change-requests/<item_id>/approve", methods=["POST"])
def approve_change_request(item_id):
    return _set_workitem_status(item_id, "Approved", expected_type="CR")


@app.route("/api/change-requests/<item_id>/reject", methods=["POST"])
def reject_change_request(item_id):
    return _set_workitem_status(item_id, "Rejected", expected_type="CR")


# ---------------------------------------------------------------------------
# FINANCIAL / BILLING -- read-only from their report sheets
# ---------------------------------------------------------------------------

def _billing_model(project_type):
    if project_type == "One Time - Implementation":
        return "Fixed Bid (Milestones)"
    if project_type == "Fixed Monthly Block of Hours":
        return "Fixed Monthly (Team Price)"
    return "T&M (Hours × Rate)"


def _all_work_items(wb):
    wi = wb[SHEET_WI]
    wih = headers_of(wi, WI_HEADER_ROW)
    return [row_to_dict(wi, r, wih, WI_FIELDS)
            for r in range(WI_FIRST_DATA_ROW, last_data_row(wi, WI_FIRST_DATA_ROW) + 1)]


def _financial_row(p, work_items):
    """Mirrors Financial Tracker's own formulas in Python -- reading that
    sheet's cells directly returns raw formula text (openpyxl can't
    evaluate formulas, and strips cached results whenever we save), so
    every number here is computed straight from the source sheets instead,
    the same approach /api/dashboard already uses. Genuinely manual-input
    columns on that sheet (Invoiced/Received/SOW Pending/Blended Rate)
    stay None -- there's no source data for them to compute from, and the
    frontend renders that as "--" rather than a fabricated 0.
    """
    actual_hrs = sum(
        (i.get("actual_effort") or 0) for i in work_items
        if i.get("type") == "Task" and i.get("module") == p.get("project_name")
    )
    planned_hrs = sum(
        (i.get("original_effort") or 0) for i in work_items
        if i.get("type") == "Task" and i.get("module") == p.get("project_name")
    )
    sow_hrs = p.get("sow_hrs") or 0
    sow_amount = p.get("sow_value")
    blended_rate = None  # manual-input on the sheet; no source to derive it from
    budget_utilization = round((actual_hrs / sow_hrs) * 100, 1) if sow_hrs else 0
    ev = (sow_hrs - actual_hrs) * (blended_rate or 0)
    return {
        # -- Project Details --
        "project_id": p.get("project_id"),
        "project_name": p.get("project_name"),
        "pm": p.get("pm"),
        "account_manager": p.get("account_manager"),
        "project_type": p.get("project_type"),
        "status": p.get("status"),
        "start_date": p.get("start_date"),
        "end_date": p.get("end_date"),
        "budget": sow_amount,           # no separate "Budget" column exists; SOW Value is the project's budget
        "sow_amount": sow_amount,
        # -- Financial Details --
        "actual_hours": actual_hrs,
        "planned_hours": planned_hrs,
        "blended_rate": blended_rate,
        "revenue": None,                # not tracked anywhere in the workbook
        "received": None,               # manual-input on Financial Tracker
        "outstanding": None,            # manual-input on Financial Tracker (AR Outstanding)
        "earned_value": ev,
        "budget_utilization": budget_utilization,
        "invoice_amount": None,         # manual-input on Financial Tracker (Invoiced)
        "margin": None,                 # would need resource cost data not present in the workbook
        "forecast": None,               # not tracked anywhere in the workbook
    }


def _milestones_for_project(wb, project_id, sow_value):
    """The Fixed Bid milestone-billing block on Billing Milestones is a
    per-project table (Project / Milestone / % of SOW / Amount / Status)
    located dynamically by its own header row, not a fixed range -- so
    this still works if rows get added/removed above it."""
    bm = wb["Billing Milestones"]
    header_row = None
    for r in range(1, bm.max_row + 1):
        vals = [bm.cell(row=r, column=c).value for c in range(1, 6)]
        if vals[:2] == ["Project", "Milestone"]:
            header_row = r
            break
    if not header_row:
        return []
    milestones = []
    r = header_row + 1
    while r <= bm.max_row:
        pid = bm.cell(row=r, column=1).value
        if not pid or not re.match(r"^P\d+$", str(pid)):
            break
        if pid == project_id:
            pct = bm.cell(row=r, column=3).value or 0
            milestones.append({
                "milestone": bm.cell(row=r, column=2).value,
                "pct_of_sow": pct,
                "amount": round((sow_value or 0) * pct, 2),
                "status": bm.cell(row=r, column=5).value,
            })
        r += 1
    return milestones


def _billing_row(wb, p):
    sow_value = p.get("sow_value")
    return {
        # -- Project Details --
        "project_id": p.get("project_id"),
        "project_name": p.get("project_name"),
        "pm": p.get("pm"),
        "project_type": p.get("project_type"),
        "status": p.get("status"),
        "billing_model": _billing_model(p.get("project_type")),
        "sow_value": sow_value,
        # -- Invoice details (no dedicated invoicing sheet exists in the
        # workbook yet, so these stay manual-input placeholders rather
        # than fabricated numbers) --
        "invoice_date": None,
        "due_date": None,
        "paid_date": None,
        "invoice_amount": None,
        "payment_status": None,
        "outstanding_amount": None,
        # -- Milestones (Fixed Bid projects only) --
        "milestones": _milestones_for_project(wb, p.get("project_id"), sow_value)
                      if p.get("project_type") == "One Time - Implementation" else [],
    }


@app.route("/api/financial")
def financial():
    wb = load_wb()
    pm = wb[SHEET_PM]
    pmh = headers_of(pm, PM_HEADER_ROW)
    work_items = _all_work_items(wb)

    rows = []
    for r in range(PM_FIRST_DATA_ROW, last_data_row(pm, PM_FIRST_DATA_ROW) + 1):
        p = row_to_dict(pm, r, pmh, PM_FIELDS)
        if not p.get("project_id"):
            continue
        rows.append(_financial_row(p, work_items))
    return jsonify({"items": rows})


@app.route("/api/billing")
def billing():
    wb = load_wb()
    pm = wb[SHEET_PM]
    pmh = headers_of(pm, PM_HEADER_ROW)

    rows = []
    for r in range(PM_FIRST_DATA_ROW, last_data_row(pm, PM_FIRST_DATA_ROW) + 1):
        p = row_to_dict(pm, r, pmh, PM_FIELDS)
        if not p.get("project_id"):
            continue
        rows.append(_billing_row(wb, p))
    return jsonify({"items": rows})


# ---------------------------------------------------------------------------
# SEARCH
# ---------------------------------------------------------------------------

@app.route("/api/search")
def search():
    q = (request.args.get("q") or "").lower()
    if not q:
        return jsonify({"projects": [], "resources": [], "workitems": []})
    wb = load_wb()
    pm, res, wi = wb[SHEET_PM], wb[SHEET_RES], wb[SHEET_WI]
    pmh, resh, wih = headers_of(pm, PM_HEADER_ROW), headers_of(res, RES_HEADER_ROW), headers_of(wi, WI_HEADER_ROW)

    projects = []
    for r in range(PM_FIRST_DATA_ROW, last_data_row(pm, PM_FIRST_DATA_ROW) + 1):
        d = row_to_dict(pm, r, pmh, PM_FIELDS)
        if d.get("project_id") and q in " ".join(str(v).lower() for v in d.values() if v):
            projects.append(d)

    resources = []
    for r in range(RES_FIRST_DATA_ROW, last_data_row(res, RES_FIRST_DATA_ROW, skip_prefix="note") + 1):
        d = row_to_dict(res, r, resh, RES_FIELDS)
        if d.get("employee") and q in " ".join(str(v).lower() for v in d.values() if v):
            resources.append(d)

    items = []
    for r in range(WI_FIRST_DATA_ROW, last_data_row(wi, WI_FIRST_DATA_ROW) + 1):
        d = row_to_dict(wi, r, wih, WI_FIELDS)
        if d.get("id") and q in " ".join(str(v).lower() for v in d.values() if v):
            items.append(d)

    return jsonify({"projects": projects[:10], "resources": resources[:10], "workitems": items[:10]})


# ---------------------------------------------------------------------------
# LOOKUPS (dropdown sources for forms)
# ---------------------------------------------------------------------------

@app.route("/api/lookups")
def lookups():
    wb = load_wb()
    pm, res, wi = wb[SHEET_PM], wb[SHEET_RES], wb[SHEET_WI]
    pmh, resh, wih = headers_of(pm, PM_HEADER_ROW), headers_of(res, RES_HEADER_ROW), headers_of(wi, WI_HEADER_ROW)

    projects = []
    for r in range(PM_FIRST_DATA_ROW, last_data_row(pm, PM_FIRST_DATA_ROW) + 1):
        name = pm.cell(row=r, column=pmh.get("Project Name", 2)).value
        if name:
            projects.append({"id": pm.cell(row=r, column=pmh.get("Project ID", 1)).value, "name": name})

    resources = []
    for r in range(RES_FIRST_DATA_ROW, last_data_row(res, RES_FIRST_DATA_ROW, skip_prefix="note") + 1):
        emp = res.cell(row=r, column=1).value
        if emp:
            resources.append(row_to_dict(res, r, resh, RES_FIELDS))

    def unique_col(header_name):
        col = wih.get(header_name)
        if not col:
            return []
        vals = set()
        for r in range(WI_FIRST_DATA_ROW, last_data_row(wi, WI_FIRST_DATA_ROW) + 1):
            v = wi.cell(row=r, column=col).value
            if v:
                vals.add(v)
        return sorted(vals)

    return jsonify({
        "projects": projects,
        "resources": resources,
        "managers": sorted({r["manager"] for r in resources if r.get("manager") and r["manager"] != "-"}),
        "priorities": unique_col("Priority") or ["P1", "P2", "P3"],
        "statuses": unique_col("Status") or ["To Do", "In Progress", "Completed", "Blocked"],
        "categories": unique_col("Category"),
        "event_bug_types": unique_col("Event / BugType"),
        "activities": unique_col("Activity"),
        "sub_activities": unique_col("Sub-Activity"),
        "domains": sorted({p for p in (pm.cell(row=r, column=pmh.get("Domain", 10)).value
                                        for r in range(PM_FIRST_DATA_ROW, last_data_row(pm, PM_FIRST_DATA_ROW) + 1)) if p}),
    })


# ---------------------------------------------------------------------------
# REPORTS / EXPORT
# ---------------------------------------------------------------------------

@app.route("/api/export")
def export_workbook():
    return send_file(XLSX_PATH, as_attachment=True, download_name="OrbitFlow_PM_Tracker_v3.xlsx")


@app.route("/api/reports/<report_type>")
def download_report(report_type):
    wb = load_wb()
    out_wb = openpyxl.Workbook()
    out_ws = out_wb.active

    sheet_map = {
        "project": (SHEET_PM, PM_HEADER_ROW, PM_FIRST_DATA_ROW, PM_FIELDS),
        "resource": (SHEET_RES, RES_HEADER_ROW, RES_FIRST_DATA_ROW, RES_FIELDS),
        "task": (SHEET_WI, WI_HEADER_ROW, WI_FIRST_DATA_ROW, WI_FIELDS),
        "risk": (SHEET_WI, WI_HEADER_ROW, WI_FIRST_DATA_ROW, WI_FIELDS),
        "financial": ("Financial Tracker", 4, 5, None),
    }
    if report_type not in sheet_map:
        return jsonify({"error": "Unknown report type"}), 404

    sheet_name, header_row, first_row, field_map = sheet_map[report_type]
    ws = wb[sheet_name]
    headers = headers_of(ws, header_row)
    if field_map:
        cols = list(field_map.values())
    else:
        cols = list(headers.keys())
    out_ws.append(cols)
    for r in range(first_row, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=headers[c]).value if headers.get(c) else None for c in cols]
        if any(v is not None for v in row_vals):
            if report_type == "risk" and row_vals[cols.index("Type")] != "Risk":
                continue
            if report_type == "task" and row_vals[cols.index("Type")] not in ("Task", "Bug"):
                continue
            out_ws.append(row_vals)

    buf = io.BytesIO()
    out_wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True, download_name=f"OrbitFlow_{report_type}_report.xlsx")


# ---------------------------------------------------------------------------
# SETTINGS
# ---------------------------------------------------------------------------

@app.route("/api/settings")
def settings():
    wb = load_wb()
    return jsonify({
        "workbook_path": str(XLSX_PATH),
        "sheets": wb.sheetnames,
        "last_modified": datetime.datetime.fromtimestamp(XLSX_PATH.stat().st_mtime).isoformat(),
        "size_kb": round(XLSX_PATH.stat().st_size / 1024, 1),
    })


@app.route("/")
def index():
    return send_from_directory(app.static_folder, "index.html")


if __name__ == "__main__":
    print(f"Using workbook: {XLSX_PATH}")
    debug = os.environ.get("FLASK_DEBUG", "0") == "1"
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=debug, host="0.0.0.0", port=port)
